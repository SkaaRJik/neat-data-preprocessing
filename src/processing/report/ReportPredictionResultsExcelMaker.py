import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.exceptions.WrongNormalizationMethodException import WrongNormalizationMethodException
from src.processing.normalization.dataset_normalizer import DatasetNormalizer
from src.processing.normalization.dataset_normalizer_factory import DatasetNormalizerFactory


class ReportPredictionResultsExcelMaker(object):
    colors = [
        'FF00FF',
        '800080',
        'FF0000',
        '800000',
        'FFFF00',
        '808000',
        '00FF00',
        '008000',
        '00FFFF',
        '008080',
        '0000FF',
        '000080',
        '000000',
        '808080',
        'C0C0C0'
    ]

    neat_settings_translations = {
        'HEADER_GENETIC_ALGORITHM': 'Параметры генетического алгоритма',
        'HEADER_EPOCH_CONTROL': 'Контроль эпох',
        'HEADER_ACTIVATION_FUNCTIONS': 'Настройки активационных функций:',
        'HEADER_NETWORK_SETTING': 'Настройки нейросети',
        'HEADER_NICHE_SETTING': 'Настройки нишинга',
        'HEADER_SPECIES_CONTROL': 'Контроль видообразования',
        'HEADER_LIFE_CONTROL': 'Контроль жизнеспособности',
        'PROBABILITY.MUTATION': 'Вероятность мутации',
        'PROBABILITY.CROSSOVER': 'Вероятность кроссинговера',
        'PROBABILITY.ADDLINK': 'Вероятность новой связи',
        'PROBABILITY.ADDNODE': 'Вероятность нового нейрона',
        'PROBABILITY.NEWACTIVATIONFUNCTION': 'Вероятность мутации актив. функции',
        'PROBABILITY.MUTATEBIAS': 'Вероятность мутации байеса',
        'PROBABILITY.TOGGLELINK': 'Вероятность включ/исключ. связи',
        'PROBABILITY.WEIGHT.REPLACED': 'Вероятность изменения веса',
        'GENERATOR.SEED': 'Семя генератора случ. чисел',
        'EXCESS.COEFFICIENT': 'Коэф-нт дополнительных генов (c1)',
        'DISJOINT.COEFFICIENT': 'Коэф-нт непересекающихся генов (c2)',
        'WEIGHT.COEFFICIENT': 'Коэф-нт средней разницы весов совпадающих генов (c3)',
        'COMPATABILITY.THRESHOLD': 'Пороговое значение соответствия особей внутри вида',
        'COMPATABILITY.CHANGE': 'Значение совместимости мутаций',
        'SPECIE.COUNT': 'Макс. количество видов',
        'SURVIVAL.THRESHOLD': 'Пороговое значение выживания',
        'SPECIE.AGE.THRESHOLD': 'Макс. время жизни особи',
        'SPECIE.YOUTH.THRESHOLD': 'Порог, до которого особь считается молодой',
        'SPECIE.OLD.PENALTY': 'Штраф для старой особи',
        'SPECIE.YOUTH.BOOST': 'Ускорение для молодой особи',
        'SPECIE.FITNESS.MAX': 'Максимальное значение приспособленности',
        'MAX.PERTURB': 'Макс. смащение весов',
        'MAX.BIAS.PERTURB': 'Макс. смещение байеса',
        'FEATURE.SELECTION': 'Отбор фич',
        'RECURRENCY.ALLOWED': 'Разрешить рекурентность',
        'INPUT.ACTIVATIONFUNCTIONS': 'Входной слой',
        'OUTPUT.ACTIVATIONFUNCTIONS': 'Выходной слой',
        'HIDDEN.ACTIVATIONFUNCTIONS': 'Скрытый слой',
        'ELE.EVENTS': 'Включить контроль вымирания',
        'ELE.SURVIVAL.COUNT': 'Extinction survival count',
        'ELE.EVENT.TIME': 'Время вымирания',
        'KEEP.BEST.EVER': 'Всегда сохранять лучшего',
        'EXTRA.FEATURE.COUNT': 'Дополнительные аллели',
        'POP.SIZE': 'Размер популяции',
        'NUMBER.EPOCHS': 'Количество эпох',
        'org.neat4j.neat.nn.core.functions.LinearFunction': 'y = x',
        'org.neat4j.neat.nn.core.functions.SigmoidFunction': 'sigmoid(x)',
        'org.neat4j.neat.nn.core.functions.TanhFunction': 'tg(x)',
        'TERMINATION.VALUE.TOGGLE': 'Включить прерывание',
        'TERMINATION.VALUE': 'Значение остановки',
        '': '',
        'minMax': 'Минимаксное скалирование'
    }

    def __init__(self):
        self._dataset_normalizer_factory = DatasetNormalizerFactory()

    def makeExcelReport(self, source_file: str,
                        result_file: str,
                        train_errors: list,
                        test_errors: list,
                        prediction_error,
                        window_train_statistic: dict,
                        columns: list,
                        legend: dict,
                        neat_settings: list,
                        normalization_method: str,
                        enable_log_transform: bool,
                        prediction_period: int,
                        prediction_window_size: int,
                        train_end_index: int,
                        test_end_index: int,
                        normalization_statistic: dict
                        ) -> (str, dict):

        input_columns = []
        output_columns = []

        for column in columns:  # type: dict
            if column.get('columnType') == 'Output':
                output_columns.append(column.get("columnName"))
            if column.get('columnType') == 'Input':
                input_columns.append(column.get("columnName"))

        source_data_frame: pd.DataFrame = pd.read_csv(source_file, sep=';')
        prediction_data_frame: pd.DataFrame = pd.read_csv(result_file, sep=';')

        for column_name in output_columns:
            column_name_with_predicted_values = f'{column_name}.1'
            source_data_frame.insert(source_data_frame.columns.size, column_name_with_predicted_values,
                                     source_data_frame[column_name])

        prediction_data_frame: pd.DataFrame = self._denormalize_data(normalization_method, enable_log_transform,
                                                                     prediction_data_frame, source_data_frame)
        wb = Workbook()
        # regex = re.compile(r'[^\d\wА-Яа-я]', re.IGNORECASE)

        new_legend_data: list = self._calculate_legend(legend.get('data'), legend.get('increment').get('increment'),
                                                       prediction_period)
        self._fill_prediction_sheet(legend=legend, new_legend_data=new_legend_data, output_columns=output_columns,
                                    prediction_data_frame=prediction_data_frame, prediction_error=prediction_error,
                                    source_data_frame=source_data_frame, window_size=prediction_window_size,
                                    prediction_period=prediction_period, wb=wb)

        self._fill_train_sheet(test_errors=test_errors, train_errors=train_errors, wb=wb, neat_settings=neat_settings,
                               source_dataframe=source_data_frame, train_end_index=train_end_index,
                               test_end_index=test_end_index, legend=legend,
                               normalization_statistic=normalization_statistic,
                               normalization_method=normalization_method)

        self._fill_factor_signs_sheet(legend=legend, new_legend_data=new_legend_data, output_columns=output_columns,
                                      prediction_data_frame=prediction_data_frame, prediction_error=prediction_error,
                                      source_data_frame=source_data_frame, window_size=prediction_window_size,
                                      prediction_period=prediction_period, wb=wb,
                                      window_train_statistic=window_train_statistic)

        prediction_result = {
            'legend_label': legend.get('header'),
            'legend': new_legend_data,
            'values': []
        }

        for column_name in output_columns:
            column_name_with_predicted_values = f'{column_name}.1'
            prediction_result['values'].append({
                'name': column_name,
                'fact_values': source_data_frame[column_name_with_predicted_values].to_list(),
                'predicted_values': prediction_data_frame[column_name_with_predicted_values].to_list(),
            })





        report_filename = '/tmp/report-1.xlsx'
        wb.save(report_filename)
        return report_filename, prediction_result

    def _fill_prediction_sheet(self, legend: dict, new_legend_data: list, output_columns: list,
                               prediction_data_frame: pd.DataFrame, prediction_error: float,
                               source_data_frame: pd.DataFrame, window_size: int, prediction_period: int, wb):
        i = 0
        column_offset = 13
        sheet = wb.active
        sheet.title = 'Прогноз'
        for column_name in output_columns:  # type: str
            # sheet_name = re.sub(regex, '', column_name.strip().replace(" ", "_"))

            sheet.cell(1, (i * column_offset) + 1, 'Прогноз целевого показателя:')
            sheet.cell(1, (i * column_offset) + 2, column_name)
            sheet.cell(2, (i * column_offset) + 1, 'Ошибка прогнозирования:')
            sheet.cell(2, (i * column_offset) + 2, prediction_error)
            sheet.cell(3, (i * column_offset) + 1, 'Размер окна')
            sheet.cell(3, (i * column_offset) + 2, window_size)
            sheet.cell(4, (i * column_offset) + 1, 'Период рогнозирования:')
            sheet.cell(4, (i * column_offset) + 2, prediction_period)

            legend_values_cell = sheet.cell(6, (i * column_offset) + 1, legend.get("header"))
            fact_values_cell = sheet.cell(6, (i * column_offset) + 2, 'Факт')
            predict_values_cell = sheet.cell(6, (i * column_offset) + 3, 'Прогноз')

            column_name_with_predicted_values = f'{column_name}.1'
            legend_end_address = self._write_vector_as_column(column_to_paste=(i * column_offset) + 1,
                                                              row_to_paste=7,
                                                              sheet=sheet,
                                                              vector=new_legend_data)
            self._write_vector_as_column(column_to_paste=(i * column_offset) + 2,
                                         row_to_paste=7,
                                         sheet=sheet,
                                         vector=source_data_frame[column_name].values)
            predicted_end_address = self._write_vector_as_column(column_to_paste=(i * column_offset) + 3,
                                                                 row_to_paste=7,
                                                                 sheet=sheet,
                                                                 vector=prediction_data_frame[
                                                                     column_name_with_predicted_values].values)

            self._create_chart_for_factor_and_predicted_values(sheet=sheet,
                                                               title=column_name,
                                                               legend_name=legend.get("header"),
                                                               chart_pos=sheet.cell(1,
                                                                                    (i * column_offset) + 4).coordinate,
                                                               data_reference=Reference(sheet,
                                                                                        min_col=fact_values_cell.column,
                                                                                        min_row=fact_values_cell.row,
                                                                                        max_col=predict_values_cell.column,
                                                                                        max_row=predicted_end_address))
            i += 1

    def _fill_train_sheet(self, test_errors: list, train_errors: list, wb: Workbook, neat_settings: list,
                          source_dataframe: pd.DataFrame, train_end_index: int, test_end_index: int, legend: dict,
                          normalization_statistic: dict, normalization_method: str):
        column_offset = 4
        wb.create_sheet('Обучение')
        sheet = wb['Обучение']
        sheet.cell(1, (0 * column_offset) + 1, "Ошибка обучения")
        sheet.cell(1, (0 * column_offset) + 2, train_errors[-1])
        sheet.cell(1, (0 * column_offset) + 3, "Ошибка тестирования")
        sheet.cell(1, (0 * column_offset) + 4, test_errors[-1])

        cell = sheet.cell(3, 1, 'Обучение')
        cell.fill = PatternFill(start_color='00EE00', end_color='00EE00', fill_type="solid")

        cell = sheet.cell(3, 2, 'Тестирование')
        cell.fill = PatternFill(start_color='EEEE00', end_color='EEEE00', fill_type="solid")

        sheet.cell(4, 1, legend.get("header"))
        self._write_vector_as_column(column_to_paste=1,
                                     row_to_paste=5,
                                     sheet=sheet,
                                     vector=legend.get("data"),
                                     train_end_index=train_end_index,
                                     test_end_index=test_end_index)

        next_free_column = 2
        for i, column_name in enumerate(source_dataframe.columns):
            sheet.cell(4, next_free_column, column_name)
            self._write_vector_as_column(column_to_paste=next_free_column,
                                         row_to_paste=5,
                                         sheet=sheet,
                                         vector=source_dataframe[column_name].values,
                                         train_end_index=train_end_index,
                                         test_end_index=test_end_index)
            next_free_column = i + 2

        epochs = x = [i for i in range(1, len(train_errors) + 1)]

        sheet.cell(4, next_free_column + (0 * column_offset) + 1, "Эпоха")
        train_errors_cell = sheet.cell(4, next_free_column + (0 * column_offset) + 2, 'Ошибка обучения')
        self._write_vector_as_column(column_to_paste=next_free_column + (0 * column_offset) + 1,
                                     row_to_paste=5,
                                     sheet=sheet,
                                     vector=epochs)
        train_end_row = self._write_vector_as_column(column_to_paste=next_free_column + (0 * column_offset) + 2,
                                                     row_to_paste=5,
                                                     sheet=sheet,
                                                     vector=train_errors)

        sheet.cell(4, next_free_column + (1 * column_offset) + 1, "Эпоха")
        test_errors_cell = sheet.cell(4, next_free_column + (1 * column_offset) + 2, 'Ошибка тестирования')
        self._write_vector_as_column(column_to_paste=next_free_column + (1 * column_offset) + 1,
                                     row_to_paste=5,
                                     sheet=sheet,
                                     vector=epochs)
        test_end_row = self._write_vector_as_column(column_to_paste=next_free_column + (1 * column_offset) + 2,
                                                    row_to_paste=5,
                                                    sheet=sheet,
                                                    vector=test_errors)
        self._create_chart_for_factor_and_predicted_values(sheet=sheet,
                                                           title='Ошибка обучения',
                                                           legend_name='Эпоха',
                                                           chart_pos=sheet.cell(1, next_free_column + (
                                                                   1 * column_offset) + column_offset).coordinate,
                                                           data_reference=Reference(sheet,
                                                                                    min_col=train_errors_cell.column,
                                                                                    min_row=train_errors_cell.row,
                                                                                    max_col=train_errors_cell.column,
                                                                                    max_row=train_end_row),
                                                           marker='none'
                                                           )
        self._create_chart_for_factor_and_predicted_values(sheet=sheet,
                                                           title='Ошибка тестирования',
                                                           legend_name='Эпоха',
                                                           chart_pos=sheet.cell(16, next_free_column + (
                                                                   1 * column_offset) + column_offset).coordinate,
                                                           data_reference=Reference(sheet,
                                                                                    min_col=test_errors_cell.column,
                                                                                    min_row=test_errors_cell.row,
                                                                                    max_col=test_errors_cell.column,
                                                                                    max_row=test_end_row),
                                                           marker='none'
                                                           )

        next_free_row, next_free_column = self._fill_neat_settings(sheet=sheet, col_to_paste=next_free_column + (1 * column_offset) + column_offset,
                                 row_to_paste=32, neat_settings=neat_settings)
        next_free_column += 1

        sheet.cell(32, next_free_column + (1 * column_offset) + 1, "Нормализация:")
        method = self.neat_settings_translations.get(normalization_method)
        method = method if method is not None else normalization_method
        sheet.cell(32, next_free_column + (1 * column_offset) + 2, method)
        sheet.cell(33, next_free_column + (1 * column_offset) + 1, 'Статистика распределения')
        category_cell: Cell = sheet.cell(34, next_free_column + (1 * column_offset) + 1, 'Период')
        values_cell: Cell = sheet.cell(34, next_free_column + (1 * column_offset) + 2, 'Процент нормализованных значений в периоде')
        category_end_index = self._write_vector_as_column(column_to_paste= next_free_column + (1 * column_offset) + 1,  row_to_paste=35,sheet=sheet,vector=list(normalization_statistic.keys()))
        values_end_index = self._write_vector_as_column(column_to_paste= next_free_column + (1 * column_offset) + 2,  row_to_paste=35,sheet=sheet,vector=list(normalization_statistic.values()))

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = 'Распределение нормализованных значений'
        chart1.y_axis.title = 'Процент значений содержащихся в периоде'
        chart1.x_axis.title = 'Период'

        data = Reference(sheet, min_col=values_cell.column, min_row=values_cell.row, max_row=values_end_index, max_col=values_cell.column)
        cats = Reference(sheet, min_col=category_cell.column, min_row=category_cell.row+1, max_row=category_end_index, max_col=category_cell.column)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        sheet.add_chart(chart1, sheet.cell(34, next_free_column + (1 * column_offset) + 3).coordinate)







    def _fill_neat_settings(self, sheet: Worksheet, col_to_paste: int, row_to_paste: int, neat_settings: list) -> tuple:
        next_row = row_to_paste
        section: dict
        end_column_index = col_to_paste
        for section in neat_settings:
            should_render = section.get('show')
            if not should_render:
                continue
            translate: str = ReportPredictionResultsExcelMaker.neat_settings_translations[section.get('header')]
            sheet.cell(row=next_row, column=col_to_paste, value=translate)
            next_row += 1
            section_params = section.get('params')
            param: dict
            for param in section_params:
                should_render = param.get('showInGui')
                if not should_render:
                    continue
                translate: str = ReportPredictionResultsExcelMaker.neat_settings_translations[param.get('name')]
                sheet.cell(row=next_row, column=col_to_paste, value=translate)
                value = param.get('value')
                if type(value) is list:
                    for i in range(len(value)):
                        end_column_index = max(end_column_index, col_to_paste + i + 1)
                        sheet.cell(row=next_row, column=col_to_paste + i + 1,
                                   value=ReportPredictionResultsExcelMaker.neat_settings_translations[value[i]])
                elif type(value) is bool:
                    sheet.cell(row=next_row, column=col_to_paste + 1, value='Да' if value is True else 'Нет')
                else:
                    sheet.cell(row=next_row, column=col_to_paste + 1, value=value)
                next_row += 1

        return next_row, end_column_index

    def _fill_factor_signs_sheet(self, wb: Workbook, legend: dict, new_legend_data: list, output_columns: list,
                                 prediction_data_frame: pd.DataFrame, prediction_error: float,
                                 source_data_frame: pd.DataFrame, window_size: int, prediction_period: int,
                                 window_train_statistic: dict):
        wb.create_sheet('Факторные_признаки')
        sheet = wb['Факторные_признаки']
        column_offset = 16
        factor_signs: list = window_train_statistic.get('factorSigns')
        factor_sign: dict
        for i, factor_sign in enumerate(factor_signs):
            column_name = factor_sign.get('name')
            sheet.cell(row=1, column=i * column_offset + 1, value='Показатель')
            sheet.cell(row=1, column=i * column_offset + 2, value=column_name)
            sheet.cell(row=2, column=i * column_offset + 1, value='Ошибка обучения')
            sheet.cell(row=2, column=i * column_offset + 2, value=factor_sign.get('trainError'))
            sheet.cell(row=3, column=i * column_offset + 1, value='Ошибка тестирования')
            sheet.cell(row=3, column=i * column_offset + 2, value=factor_sign.get('testError'))
            sheet.cell(row=5, column=i * column_offset + 1, value=legend.get('header'))
            fact_cell = sheet.cell(row=5, column=i * column_offset + 2, value='Факт')
            predict_cell = sheet.cell(row=5, column=i * column_offset + 3, value='Прогноз')

            self._write_vector_as_column(row_to_paste=6, column_to_paste=i * column_offset + 1, sheet=sheet,
                                         vector=new_legend_data)
            self._write_vector_as_column(row_to_paste=6, column_to_paste=i * column_offset + 2, sheet=sheet,
                                         vector=source_data_frame[column_name].values)
            predict_end_row = self._write_vector_as_column(row_to_paste=6, column_to_paste=i * column_offset + 3,
                                                           sheet=sheet,
                                                           vector=prediction_data_frame[column_name].values)
            self._create_chart_for_factor_and_predicted_values(sheet=sheet,
                                                               title=column_name,
                                                               legend_name=legend.get("header"),
                                                               chart_pos=sheet.cell(1,
                                                                                    i * column_offset + 4).coordinate,

                                                               data_reference=Reference(sheet,
                                                                                        min_col=fact_cell.column,
                                                                                        min_row=fact_cell.row,
                                                                                        max_col=predict_cell.column,
                                                                                        max_row=predict_end_row),
                                                               )

    def _write_vector_as_column(self, column_to_paste: int, row_to_paste: int, sheet, vector: list,
                                train_end_index: int = -1, test_end_index: int = -1) -> int:
        end_address = row_to_paste
        for i, value in enumerate(vector):
            cell = sheet.cell(end_address, column_to_paste, value)

            if (i <= test_end_index):
                cell.fill = PatternFill(start_color='EEEE00', end_color='EEEE00', fill_type="solid")

            if (i <= train_end_index):
                cell.fill = PatternFill(start_color='00EE00', end_color='00EE00', fill_type="solid")

            end_address = row_to_paste + 1 + i
        return end_address

    def _denormalize_data(self, normalization_method: str, log_transformed: bool,
                          dataframe_to_denormalize: pd.DataFrame, source_data_frame: pd.DataFrame) -> pd.DataFrame:
        data_normalizer: DatasetNormalizer = self._dataset_normalizer_factory.getNormalizer(normalization_method)

        if data_normalizer is None:
            raise WrongNormalizationMethodException('{0} - method not found'.format(normalization_method))

        return data_normalizer.denormalize(dataframe_to_denormalize, source_data_frame, log_transformed)

    def _calculate_legend(self, legend_data: list, increment, prediction_period: int) -> list:
        new_legend_data = legend_data[:]

        for i in range(prediction_period):
            new_legend_data.append(new_legend_data[-1] + increment)

        return new_legend_data

    def _create_chart_for_factor_and_predicted_values(self, sheet, title, legend_name, chart_pos: str,
                                                      data_reference: Reference,
                                                      marker='circle'):
        chart = LineChart()
        chart.title = title
        chart.style = 14

        chart.y_axis.title = 'Значения'
        chart.x_axis.title = legend_name
        chart.x_axis.tickLblPos = "low"

        chart.add_data(data_reference, titles_from_data=True)
        for i, series in enumerate(chart.series):
            series.marker.symbol = marker
            series.graphicalProperties.line.solidFill = self.colors[i % len(self.colors)]
            series.marker.graphicalProperties.solidFill = self.colors[i % len(self.colors)]
            series.marker.graphicalProperties.line.solidFill = self.colors[i % len(self.colors)]

        sheet.add_chart(chart, chart_pos)
